"""Safely append one structured paper-reading row to the review workbook."""

from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import unicodedata
import uuid
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import TableColumn


WORKBOOK_ENV_VAR = "PAPER_READING_WORKBOOK"

FIELDS = [
    "序号",
    "论文名称",
    "作者",
    "期刊",
    "期刊等级",
    "发表时间",
    "关键词",
    "研究问题",
    "研究结论",
    "研究逻辑",
    "实证模型",
    "数据来源和变量设置",
    "可进一步延伸的研究设计",
]


def workbook_from_environment() -> Path | None:
    value = os.environ.get(WORKBOOK_ENV_VAR, "").strip()
    return Path(value).expanduser() if value else None


def copy_cell_format(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def normalize_key(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    return re.sub(r"[\s《》〈〉“”‘’'\"·—–_\-:：;；,，.。()（）\[\]【】]+", "", text)


def sheet_snapshot(ws, excluded_rows: set[int] | None = None) -> dict:
    excluded_rows = excluded_rows or set()
    cells = {}
    for row in range(1, ws.max_row + 1):
        if row in excluded_rows:
            continue
        for col in range(1, min(ws.max_column, 12) + 1):
            cell = ws.cell(row, col)
            cells[(row, col)] = (
                cell.value,
                cell.style_id,
                cell.number_format,
                cell.alignment.horizontal,
                cell.alignment.vertical,
                cell.alignment.wrap_text,
            )
    return {
        "cells": cells,
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "merged_ranges": sorted(str(item) for item in ws.merged_cells.ranges),
        "table_names": sorted(ws.tables.keys()),
    }


def extend_table_to_thirteenth_field(table, max_row: int) -> bool:
    min_col, min_row, max_col, current_max_row = range_boundaries(table.ref)
    if min_row != 1 or min_col != 1 or max_col not in (12, 13):
        return False

    original_ref = table.ref
    original_filter_ref = table.autoFilter.ref if table.autoFilter is not None else None

    if max_col == 12:
        if len(table.tableColumns) != 12:
            raise ValueError(
                f"表格“{table.name}”的范围为12列，但列定义数量为"
                f"{len(table.tableColumns)}；已停止写入。"
            )
        next_id = max(column.id for column in table.tableColumns) + 1
        table.tableColumns.append(TableColumn(id=next_id, name=FIELDS[12]))
    elif len(table.tableColumns) != 13:
        raise ValueError(
            f"表格“{table.name}”的范围为13列，但列定义数量为"
            f"{len(table.tableColumns)}；已停止写入。"
        )

    new_ref = f"A1:M{max(current_max_row, max_row)}"
    table.ref = new_ref
    if table.autoFilter is not None:
        table.autoFilter.ref = new_ref
    return table.ref != original_ref or (
        table.autoFilter is not None and table.autoFilter.ref != original_filter_ref
    )


def ensure_thirteenth_field(ws) -> bool:
    actual = [ws.cell(1, col).value for col in range(1, 13)]
    if actual != FIELDS[:12]:
        raise ValueError(
            f"工作表“{ws.title}”前12列表头与技能字段不一致；为避免破坏原表，已停止写入。"
        )

    current = ws.cell(1, 13).value
    if current not in (None, "", FIELDS[12]):
        raise ValueError(
            f"工作表“{ws.title}”第13列表头为“{current}”，不是“{FIELDS[12]}”；已停止写入。"
        )

    changed = current != FIELDS[12]
    if changed:
        header = ws.cell(1, 13, FIELDS[12])
        copy_cell_format(ws.cell(1, 12), header)

        source_dimension = ws.column_dimensions["L"]
        target_dimension = ws.column_dimensions["M"]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden
        target_dimension.bestFit = source_dimension.bestFit
        target_dimension.outlineLevel = source_dimension.outlineLevel

        for row in range(2, ws.max_row + 1):
            copy_cell_format(ws.cell(row, 12), ws.cell(row, 13))

    has_m_validation = any(
        "M2:M1000" in str(item.sqref) for item in ws.data_validations.dataValidation
    )
    if not has_m_validation:
        validation = DataValidation(type="custom", formula1="TRUE", allow_blank=True)
        validation.promptTitle = "延伸研究设计"
        validation.prompt = "提出2—4项可执行设计；说明拓展问题、识别策略、数据或变量。"
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        validation.add("M2:M1000")
        changed = True

    filter_ref = ws.auto_filter.ref
    if filter_ref:
        min_col, min_row, _max_col, max_row = range_boundaries(filter_ref)
        new_filter_ref = (
            f"{get_column_letter(min_col)}{min_row}:M{max(max_row, ws.max_row)}"
        )
        if new_filter_ref != filter_ref:
            ws.auto_filter.ref = new_filter_ref
            changed = True

    for table in ws.tables.values():
        changed = extend_table_to_thirteenth_field(table, ws.max_row) or changed

    return changed


def find_last_title_row(ws) -> int:
    rows = [
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 2).value not in (None, "")
    ]
    return max(rows) if rows else 1


def find_duplicate_row(ws, payload: dict) -> int | None:
    title = normalize_key(payload.get("论文名称"))
    journal = normalize_key(payload.get("期刊"))
    for row in range(2, ws.max_row + 1):
        if normalize_key(ws.cell(row, 2).value) != title:
            continue
        if normalize_key(ws.cell(row, 4).value) == journal:
            return row
    return None


def translated_formula(
    source_value: object, source_coordinate: str, target_coordinate: str
) -> str:
    if isinstance(source_value, str) and source_value.startswith("="):
        try:
            return Translator(source_value, origin=source_coordinate).translate_formula(
                target_coordinate
            )
        except Exception:
            return source_value
    return "=ROW()-1"


def append_payload(ws, payload: dict) -> int:
    duplicate = find_duplicate_row(ws, payload)
    if duplicate is not None:
        raise FileExistsError(
            f"检测到重复论文，现有行号为 {duplicate}；未追加、未覆盖。"
        )

    last_row = find_last_title_row(ws)
    append_row = max(2, last_row + 1)
    style_row = last_row if last_row >= 2 else 2

    for col in range(1, 14):
        copy_cell_format(ws.cell(style_row, col), ws.cell(append_row, col))

    source_height = ws.row_dimensions[style_row].height
    if source_height is not None:
        ws.row_dimensions[append_row].height = source_height

    ws.cell(append_row, 1).value = translated_formula(
        ws.cell(style_row, 1).value,
        ws.cell(style_row, 1).coordinate,
        ws.cell(append_row, 1).coordinate,
    )
    for col, field in enumerate(FIELDS[1:], start=2):
        ws.cell(append_row, col).value = payload.get(field)

    if ws.auto_filter.ref:
        min_col, min_row, _max_col, max_row = range_boundaries(ws.auto_filter.ref)
        ws.auto_filter.ref = (
            f"{get_column_letter(min_col)}{min_row}:M{max(max_row, append_row)}"
        )
    for table in ws.tables.values():
        extend_table_to_thirteenth_field(table, append_row)

    return append_row


def validate_payload(payload: dict) -> None:
    missing = [field for field in FIELDS[1:] if field not in payload]
    if missing:
        raise ValueError(f"JSON缺少字段：{', '.join(missing)}")
    if not str(payload.get("论文名称") or "").strip():
        raise ValueError("“论文名称”不能为空。")


def save_validate_replace(
    path: Path,
    wb,
    before: dict,
    target_sheet: str | None,
    row: int | None,
) -> Path:
    temp_path = path.with_name(f".{path.stem}.{uuid.uuid4().hex}.tmp{path.suffix}")
    backup_path = path.with_name(
        f"{path.stem}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}{path.suffix}"
    )
    check = None
    try:
        wb.save(temp_path)
        wb.close()
        check = load_workbook(temp_path, read_only=False, data_only=False)
        if check.sheetnames != list(before):
            raise ValueError("保存后工作表名称或顺序发生变化。")

        for name, original in before.items():
            ws = check[name]
            if [ws.cell(1, col).value for col in range(1, 14)] != FIELDS:
                raise ValueError(f"保存后工作表“{name}”的13列表头校验失败。")
            excluded = {row} if name == target_sheet and row is not None else set()
            current = sheet_snapshot(ws, excluded)
            if current["freeze_panes"] != original["freeze_panes"]:
                raise ValueError(f"工作表“{name}”冻结窗格发生意外变化。")
            if current["merged_ranges"] != original["merged_ranges"]:
                raise ValueError(f"工作表“{name}”合并单元格发生意外变化。")
            if current["table_names"] != original["table_names"]:
                raise ValueError(f"工作表“{name}”表格名称发生意外变化。")
            if current["cells"] != original["cells"]:
                raise ValueError(f"工作表“{name}”原有A:L内容或样式发生意外变化。")

        if target_sheet and row is not None:
            ws = check[target_sheet]
            if ws.cell(row, 2).value in (None, "") or ws.cell(row, 13).value in (
                None,
                "",
            ):
                raise ValueError("新增行的论文名称或延伸研究设计为空。")
            if not str(ws.cell(row, 1).value).startswith("="):
                raise ValueError("新增行序号公式未保留。")

        check.close()
        check = None
        shutil.copy2(path, backup_path)
        os.replace(temp_path, path)
        return backup_path
    except Exception:
        if check is not None:
            check.close()
        if temp_path.exists():
            temp_path.unlink()
        raise


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        help=(
            "Target .xlsx workbook. If omitted, use the "
            f"{WORKBOOK_ENV_VAR} environment variable."
        ),
    )
    parser.add_argument("--sheet", choices=["中文", "英文"], default="中文")
    parser.add_argument("--data-json", type=Path)
    parser.add_argument(
        "--schema-only",
        action="store_true",
        help="Only add/validate the thirteenth field without appending a paper.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook = args.workbook or workbook_from_environment()
    if workbook is None:
        raise ValueError(
            "未指定工作簿。请传入 --workbook <path> 或设置环境变量 "
            f"{WORKBOOK_ENV_VAR}。"
        )
    path = workbook.expanduser().resolve()
    if not path.exists():
        raise FileNotFoundError(f"工作簿不存在：{path}")
    if not args.schema_only and args.data_json is None:
        raise ValueError("追加论文时必须提供 --data-json。")

    wb = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    if "中文" not in wb.sheetnames or "英文" not in wb.sheetnames:
        wb.close()
        raise ValueError("工作簿必须包含“中文”和“英文”工作表。")

    before = {name: sheet_snapshot(wb[name]) for name in wb.sheetnames}
    schema_changed = False
    for ws in wb.worksheets:
        schema_changed = ensure_thirteenth_field(ws) or schema_changed

    appended_row = None
    payload = None
    if not args.schema_only:
        payload = json.loads(args.data_json.read_text(encoding="utf-8"))
        validate_payload(payload)
        duplicate = find_duplicate_row(wb[args.sheet], payload)
        if duplicate is not None:
            wb.close()
            print(
                json.dumps(
                    {
                        "action": "duplicate_skipped",
                        "workbook": str(path),
                        "sheet": args.sheet,
                        "row": duplicate,
                        "schema_saved": False,
                    },
                    ensure_ascii=False,
                )
            )
            return 0
        appended_row = append_payload(wb[args.sheet], payload)

    if not schema_changed and appended_row is None:
        wb.close()
        print(
            json.dumps(
                {"action": "schema_already_current", "workbook": str(path)},
                ensure_ascii=False,
            )
        )
        return 0

    excluded_before = {name: before[name] for name in before}
    if appended_row is not None and appended_row <= wb[args.sheet].max_row:
        original = load_workbook(path, read_only=False, data_only=False)
        excluded_before[args.sheet] = sheet_snapshot(
            original[args.sheet], {appended_row}
        )
        original.close()

    backup = save_validate_replace(
        path,
        wb,
        excluded_before,
        args.sheet if appended_row is not None else None,
        appended_row,
    )
    wb.close()
    print(
        json.dumps(
            {
                "action": "schema_updated"
                if appended_row is None
                else "paper_appended",
                "workbook": str(path),
                "backup": str(backup),
                "sheet": args.sheet if appended_row is not None else None,
                "row": appended_row,
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(
            json.dumps({"action": "error", "message": str(exc)}, ensure_ascii=False),
            file=sys.stderr,
        )
        raise SystemExit(1)
