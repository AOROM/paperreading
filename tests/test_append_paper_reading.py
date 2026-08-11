from __future__ import annotations

import hashlib
import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from paperreading.domain import PaperRecord  # noqa: E402
from paperreading.exporters.excel import ExcelExporter  # noqa: E402

SCRIPT = (
    ROOT / "skills" / "papers-reading-skill" / "scripts" / "append_paper_reading.py"
)
FIELDS = [
    "序号",
    "论文名称",
    "作者",
    "期刊",
    "期刊等级",
    "发表时间",
    "关键词",
    "研究问题",
    "研究结论",
    "研究逻辑",
    "实证模型",
    "数据来源和变量设置",
    "可进一步延伸的研究设计",
]


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def example_payload() -> dict[str, str]:
    return {
        "论文名称": "测试论文",
        "作者": "甲，乙",
        "期刊": "测试期刊",
        "期刊等级": "",
        "发表时间": "2024",
        "关键词": "测试；识别",
        "研究问题": "1. 测试问题？",
        "研究结论": "【基准结论】测试结论。",
        "研究逻辑": "理论 → 冲击 → 结果",
        "实证模型": "【基准模型】固定效应模型。",
        "数据来源和变量设置": "【样本】测试样本。",
        "可进一步延伸的研究设计": "1. 【识别强化】构造外生冲击。",
    }


def create_workbook(path: Path, bad_header: bool = False) -> None:
    workbook = Workbook()
    workbook.active.title = "中文"
    workbook.create_sheet("英文")

    for index, worksheet in enumerate(workbook.worksheets, start=1):
        headers = FIELDS[:12].copy()
        if bad_header and worksheet.title == "英文":
            headers[0] = "错误表头"
        for column, value in enumerate(headers, start=1):
            cell = worksheet.cell(1, column, value)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")

        worksheet.append(
            [
                "=ROW()-1",
                f"已有论文{index}",
                "已有作者",
                "已有期刊",
                "",
                "2023",
                "已有关键词",
                "已有问题",
                "已有结论",
                "已有逻辑",
                "已有模型",
                "已有数据",
            ]
        )
        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = "A1:L2"
        worksheet.column_dimensions["L"].width = 42
        worksheet.add_table(Table(displayName=f"PaperTable{index}", ref="A1:L2"))

    workbook.save(path)
    workbook.close()


class AppendPaperReadingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.root = Path(self.tempdir.name)
        self.workbook = self.root / "review.xlsx"
        self.payload_path = self.root / "payload.json"
        self.payload_path.write_text(
            json.dumps(example_payload(), ensure_ascii=False), encoding="utf-8"
        )

    def tearDown(self) -> None:
        self.tempdir.cleanup()

    def run_script(
        self,
        *,
        explicit_workbook: bool = True,
        env_workbook: bool = False,
    ) -> subprocess.CompletedProcess[str]:
        command = [
            sys.executable,
            str(SCRIPT),
            "--sheet",
            "中文",
            "--data-json",
            str(self.payload_path),
        ]
        if explicit_workbook:
            command.extend(["--workbook", str(self.workbook)])

        environment = os.environ.copy()
        environment["PYTHONUTF8"] = "1"
        environment.pop("PAPER_READING_WORKBOOK", None)
        if env_workbook:
            environment["PAPER_READING_WORKBOOK"] = str(self.workbook)

        return subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            env=environment,
        )

    def run_schema_only(self) -> subprocess.CompletedProcess[str]:
        environment = os.environ.copy()
        environment["PYTHONUTF8"] = "1"
        return subprocess.run(
            [
                sys.executable,
                str(SCRIPT),
                "--workbook",
                str(self.workbook),
                "--schema-only",
            ],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            env=environment,
        )

    def test_append_upgrades_schema_and_preserves_structure(self) -> None:
        create_workbook(self.workbook)
        result = self.run_script()
        self.assertEqual(result.returncode, 0, result.stderr)
        status = json.loads(result.stdout)
        self.assertEqual(status["action"], "paper_appended")
        self.assertEqual(status["row"], 3)
        self.assertTrue(Path(status["backup"]).exists())

        workbook = load_workbook(self.workbook, data_only=False)
        try:
            for worksheet in workbook.worksheets:
                self.assertEqual(
                    [worksheet.cell(1, column).value for column in range(1, 14)],
                    FIELDS,
                )
                self.assertEqual(worksheet.freeze_panes, "B2")
                self.assertIn(
                    "M2:M1000", str(worksheet.data_validations.dataValidation[0].sqref)
                )
                table = next(iter(worksheet.tables.values()))
                expected_last_row = 3 if worksheet.title == "中文" else 2
                self.assertEqual(table.ref, f"A1:M{expected_last_row}")
                self.assertEqual(table.autoFilter.ref, table.ref)
                self.assertEqual(len(table.tableColumns), 13)
                self.assertEqual(table.tableColumns[-1].name, FIELDS[12])

            chinese = workbook["中文"]
            self.assertEqual(chinese["A3"].value, "=ROW()-1")
            self.assertEqual(chinese["B3"].value, "测试论文")
            self.assertEqual(chinese["M3"].value, example_payload()[FIELDS[12]])
            self.assertEqual(chinese.column_dimensions["M"].width, 42)
        finally:
            workbook.close()

    def test_core_excel_exporter_projects_rich_record(self) -> None:
        create_workbook(self.workbook)
        record = PaperRecord.model_validate_json(
            (ROOT / "examples" / "paper-record.example.json").read_text(
                encoding="utf-8"
            )
        )
        status = ExcelExporter().export([record], self.workbook, sheet="中文")
        self.assertEqual(status["action"], "paper_appended")
        self.assertTrue(Path(str(status["backup"])).exists())

        workbook = load_workbook(self.workbook, data_only=False)
        try:
            chinese = workbook["中文"]
            self.assertEqual(chinese["B3"].value, record.metadata.title)
            self.assertIn("【基准结论】", chinese["I3"].value)
            self.assertIn("【识别强化】", chinese["M3"].value)
        finally:
            workbook.close()

    def test_duplicate_is_a_successful_no_op(self) -> None:
        create_workbook(self.workbook)
        first = self.run_script()
        self.assertEqual(first.returncode, 0, first.stderr)
        backup_count = len(list(self.root.glob("review.backup_*.xlsx")))

        second = self.run_script()
        self.assertEqual(second.returncode, 0, second.stderr)
        status = json.loads(second.stdout)
        self.assertEqual(status["action"], "duplicate_skipped")
        self.assertEqual(status["row"], 3)
        self.assertEqual(
            len(list(self.root.glob("review.backup_*.xlsx"))), backup_count
        )

        workbook = load_workbook(self.workbook)
        try:
            self.assertEqual(workbook["中文"].max_row, 3)
        finally:
            workbook.close()

    def test_header_mismatch_leaves_source_unchanged(self) -> None:
        create_workbook(self.workbook, bad_header=True)
        before = file_hash(self.workbook)
        result = self.run_script()
        self.assertNotEqual(result.returncode, 0)
        self.assertEqual(file_hash(self.workbook), before)
        status = json.loads(result.stderr)
        self.assertEqual(status["action"], "error")
        self.assertIn("表头", status["message"])

    def test_environment_variable_can_supply_workbook(self) -> None:
        create_workbook(self.workbook)
        result = self.run_script(explicit_workbook=False, env_workbook=True)
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(json.loads(result.stdout)["action"], "paper_appended")

    def test_missing_workbook_configuration_fails_clearly(self) -> None:
        result = self.run_script(explicit_workbook=False, env_workbook=False)
        self.assertNotEqual(result.returncode, 0)
        status = json.loads(result.stderr)
        self.assertEqual(status["action"], "error")
        self.assertIn("PAPER_READING_WORKBOOK", status["message"])

    def test_schema_only_repairs_filter_and_table_ranges(self) -> None:
        create_workbook(self.workbook)
        workbook = load_workbook(self.workbook)
        try:
            for worksheet in workbook.worksheets:
                worksheet["M1"] = FIELDS[12]
                validation = DataValidation(
                    type="custom", formula1="TRUE", allow_blank=True
                )
                worksheet.add_data_validation(validation)
                validation.add("M2:M1000")
            workbook.save(self.workbook)
        finally:
            workbook.close()

        result = self.run_schema_only()
        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(json.loads(result.stdout)["action"], "schema_updated")

        workbook = load_workbook(self.workbook)
        try:
            for worksheet in workbook.worksheets:
                self.assertEqual(worksheet.auto_filter.ref, "A1:M2")
                table = next(iter(worksheet.tables.values()))
                self.assertEqual(table.ref, "A1:M2")
                self.assertEqual(table.autoFilter.ref, "A1:M2")
                self.assertEqual(len(table.tableColumns), 13)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
